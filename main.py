from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
import time
import shutil
import glob
import os
import openpyxl
from tkinter import Tk     # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename
from tkinter import messagebox as tkMessageBox
import datetime

def main():

    # Use excel to get data: Trial Collection - Result.xlsm
    Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing
    file = askopenfilename()  # show an "Open" dialog box and return the path to the selected file
    file = file.replace("/", "\\")

    print(file)

    now = datetime.datetime.now()
    now = now.strftime('%m/%d/%Y')
    print(now)

    # Open Excel
    xls = openpyxl.load_workbook(file, keep_vba=True)
    #sheet = xls.active
    sheet = xls["Result"]
    maxRow = 0

    for row in sheet:
        if not all([cell.value is None for cell in row]):
            maxRow += 1

    print(maxRow)

    # Directory where to save downloaded files
    path = os.getcwd()+"\Downloads"

    pathExist = os.path.exists(path)

    if not pathExist:
        path = os.mkdir(path)


    # Download Chrome Driver
    try:
        try:
            try:
                driver = webdriver.Chrome(ChromeDriverManager().install())
                driver.quit()
                list_of_files = glob.glob(os.environ['USERPROFILE'] + "\.wdm\drivers\**\chromedriver.exe", recursive=True)
                latest_file = max(list_of_files, key=os.path.getctime)
                time.sleep(1)
                shutil.copy2(latest_file, os.getcwd() + "\\chromedriver.exe")
                del driver
            except Exception as e:
                pass
            options = Options()
            # Auto downloading
            options.add_experimental_option("prefs", {
                "download.default_directory": path,
                "download.prompt_for_download": False,  # To auto download the file
                "download.directory_upgrade": True,
                "plugins.always_open_pdf_externally": True  # It will not show PDF directly in chrome
            })
            # Make Chrome Hide and Show
            options.headless = False
            # Hide all warnings
            options.add_experimental_option("excludeSwitches", ["enable-logging"])
            driver = webdriver.Chrome(executable_path='chromedriver.exe', options=options)

        except Exception as e:
            driver = webdriver.Chrome(ChromeDriverManager().install())
            driver.quit()
            del driver
            try:
                list_of_files = glob.glob(os.environ['USERPROFILE'] + "\.wdm\drivers\**\chromedriver.exe", recursive=True)
                latest_file = max(list_of_files, key=os.path.getctime)
                time.sleep(1)
                shutil.copy2(latest_file, os.getcwd() + "\\chromedriver.exe")
            except Exception as e:
                pass

            options = Options()
            # Make Chrome Hide and Show
            options.add_experimental_option("prefs", {
                "download.default_directory": path,
                "download.prompt_for_download": False,  # To auto download the file
                "download.directory_upgrade": True,
                "plugins.always_open_pdf_externally": True  # It will not show PDF directly in chrome
            })
            options.headless = False
            # Hide all warnings
            options.add_experimental_option("excludeSwitches", ["enable-logging"])
            driver = webdriver.Chrome(executable_path='chromedriver.exe', options=options)
    except Exception as e:
        try:
            driver.quit()
        except Exception as e:
            pass
        print(e)


    timeout = 30
    url = r'https://ecf.ca2.uscourts.gov/docs1/00209020230'

    driver.get(url)
    # Wait for browser to load page
    driver.implicitly_wait(timeout)

    # Login to site

    element_present = EC.presence_of_element_located((By.XPATH, '//*[@id="loginForm:fbtnLogin"]/span'))
    WebDriverWait(driver, timeout).until(element_present)

    userBtn = driver.find_element(By.XPATH, '//*[@id="loginForm:loginName"]')
    userBtn.send_keys("57uwr63Yv")
    pwBtn = driver.find_element(By.XPATH, '//*[@id="loginForm:password"]')
    pwBtn.send_keys("98bcnu%PL")
    time.sleep(3) # Must have LOL
    loginBtn = driver.find_element(By.XPATH, '//*[@id="loginForm:fbtnLogin"]/span')
    loginBtn.click()

    # Delete all files
    for f in os.listdir(path):
        try:
            os.remove(os.path.join(path, f))
        except:
            pass

    # Create Temporary (Backup) and Final destination folder
    try:
        os.mkdir(path +"\\Temp\\")
        #os.mkdir(path + "\\Final\\")
    except:
        pass

    print("Starting to download files.")
    with open(file) as excel:
        for i in range(2, maxRow + 1):
            for j in range(1, sheet.max_column):
                rowFilename = sheet.cell(row=i, column=8).value  # File Name
                rowLink = sheet.cell(row=i, column=10).value  # Document Link
                rowLocation = sheet.cell(row=i, column=11).value  # SF Location
                botStatus = sheet.cell(row=i, column=12)  # Bot Remarks
                rowStatus = botStatus.value
                botRunDate = sheet.cell(row=i, column=13)  # Bot Run Date
                rowDate = botRunDate.value

                #print(sheet.cell(row=maxRow, column=8).value)
            if rowStatus == "Download Success":
                continue
            else:
                fileName = rowFilename + ".pdf"
                print("Downloading", fileName)

                # Combination of the URL and filename with .pdf extension
                locPath = os.path.join(rowLocation, fileName)
                print(locPath)

                # Check if file exists
                fileExist = os.path.exists(locPath)

                if fileExist:
                    print("File already exists")
                    botStatus.value = "File already exists"
                    botRunDate.value = now
                else:

                    # Start Loop
                    driver.get(rowLink)
                    try:
                        print("first try")
                        acceptBtnPresent = EC.presence_of_element_located((By.XPATH, '/html/body/center/center[2]/form/input[9]'))
                        WebDriverWait(driver, timeout).until(acceptBtnPresent)
                        acceptBtn = driver.find_element(By.XPATH, '/html/body/center/center[2]/form/input[9]')
                        acceptBtn.click()
                    except:
                        try:
                            print("second try")
                            contBtnPresent = EC.presence_of_element_located(((By.XPATH, '/html/body/div[2]/form/p/a')))
                            WebDriverWait(driver, timeout).until(contBtnPresent)
                            contBtn = driver.find_element(By.XPATH, '/html/body/div[2]/form/p/a')
                            contBtn.click()
                            time.sleep(5)
                            # View document
                            viewBtnPresent = EC.presence_of_element_located(((By.XPATH, '/html/body/div[2]/form/input')))
                            WebDriverWait(driver, timeout).until(viewBtnPresent)
                            viewDocBtn = driver.find_element(By.XPATH, '/html/body/div[2]/form/input')
                            viewDocBtn.click()
                            #time.sleep(5)
                            try:
                                print("3rd try")
                                iframe = driver.find_element(By.XPATH, '//*[@id="cmecfMainContent"]/iframe')
                                driver.switch_to.frame(iframe)
                                submit_button = driver.find_element(By.ID, 'open-button')
                                submit_button.click()
                            except:
                                print("first pass")
                                pass
                        except:
                            print("second pass")
                            pass


                    #time.sleep(10)
                    # Renaming of file, and file transfer using Shutil
                    bool_cont = True
                    # Loop until .pdf is found
                    while bool_cont == True:
                        for f in os.listdir(path):
                            print(f)
                            if (f.lower().endswith(".pdf") or f.lower().endswith("pdf")):
                                print(f)
                                downloadedFile = os.path.join(path, f)
                                print("Moving to Temp folder:", fileName)
                                shutil.copyfile(downloadedFile, path + "\\Temp\\" + fileName)
                                print("Moving to SF Location:", fileName)
                                shutil.move(downloadedFile, locPath)
                                print("Done with", fileName)

                                print("Download Success")
                                botStatus.value = "Download Success"
                                botRunDate.value = now
                                bool_cont = False
                                break
                            #else:
                            #    bool_cont = False
                            #    break

    xls.save(file)
    xls.close()

    driver.quit()
    print("Done.")
    tkMessageBox.showinfo(title="Status", message="All files have been downloaded.")


if __name__ == "__main__":
    main()